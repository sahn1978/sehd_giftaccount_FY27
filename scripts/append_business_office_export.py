#!/usr/bin/env python3
"""
append_business_office_export.py

Turns a business-office pivot export (the same shape as the file the school
sends: one row per fund/driver, one column with the total amount as of a
date) into rows appended to data/ledger.csv, the longitudinal ledger that
powers the dashboard.

USAGE
    python3 scripts/append_business_office_export.py path/to/export.xlsx
    python3 scripts/append_business_office_export.py path/to/export.xlsx --date 2026-07-17
    python3 scripts/append_business_office_export.py path/to/export.xlsx --ledger data/ledger.csv

WHAT IT EXPECTS IN THE EXPORT FILE
    Column A: "Driver" — text like "Gift: BG001523 SOE-KIN Gift Account-403130"
              or "Program: PG007894 Signorile Research Support"
    Column B: "Total amount (M/D/YYYY)" — the running balance for that fund
              as of the date in the header. Negative = still available.
              Positive = the fund has been overspent.
    A "Grand Total" row is present and gets skipped automatically.

    If your export has more than one date column (because you pasted a few
    weeks side by side), the script processes every "Total amount (...)"
    column it finds and appends one reconciliation batch per date.

WHAT IT DOES
    - Reads every fund/driver row (skips "Grand Total")
    - Splits "Gift: BG001523 SOE-KIN Gift Account-403130" into
      fund_code = BG001523, fund_name = SOE-KIN Gift Account-403130,
      category = Gift
    - Pulls the as-of date from the column header, unless --date overrides it
    - Appends one "Balance Reconciliation" row per fund to data/ledger.csv
    - Skips a date that has already been loaded, so it's safe to re-run

This never overwrites history — it only appends new dated snapshots, which
is what makes the ledger longitudinal.
"""

import argparse
import csv
import os
import re
import sys
from datetime import datetime

import openpyxl

LEDGER_HEADER = [
    "pull_date", "transaction_date", "fund_code", "fund_name", "category",
    "entry_type", "amount", "running_balance_reported", "description", "source",
]

DATE_COL_RE = re.compile(r"total amount\s*\(([^)]+)\)", re.IGNORECASE)
DRIVER_RE = re.compile(r"^\s*(Gift|Program|Endowment)\s*:\s*([A-Za-z]{2}\d+)\s+(.*)$", re.IGNORECASE)


def parse_driver(driver_text):
    """Split a 'Gift: BG001523 SOE-KIN Gift Account-403130' string into parts."""
    m = DRIVER_RE.match(driver_text.strip())
    if m:
        category, fund_code, fund_name = m.groups()
        return fund_code.strip(), fund_name.strip(), category.strip().title()
    # Fallback: keep the whole string as the fund name if it doesn't match
    # the expected pattern, so nothing silently gets dropped.
    return driver_text.strip(), driver_text.strip(), "Other"


def parse_header_date(header_text, override_date):
    if override_date:
        return override_date
    m = DATE_COL_RE.search(str(header_text))
    if not m:
        return None
    raw = m.group(1).strip()
    # Business office exports sometimes have typos like "7//10/2026" -
    # collapse repeated slashes so the date still parses.
    raw = re.sub(r"/{2,}", "/", raw)
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def load_existing_dates(ledger_path):
    dates = set()
    if not os.path.exists(ledger_path):
        return dates
    with open(ledger_path, newline="") as f:
        for row in csv.DictReader(f):
            if row.get("entry_type") == "Balance Reconciliation":
                dates.add(row.get("pull_date"))
    return dates


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("export_file", help="Path to the business office .xlsx export")
    ap.add_argument("--ledger", default="data/ledger.csv", help="Path to the ledger CSV (default: data/ledger.csv)")
    ap.add_argument("--date", default=None, help="Override the as-of date (YYYY-MM-DD) instead of reading it from the column header")
    ap.add_argument("--sheet", default=None, help="Sheet name to read (default: first sheet)")
    ap.add_argument("--source-label", default=None, help="Text stored in the 'source' column (default: 'Business Office Export <file name>')")
    ap.add_argument("--force", action="store_true", help="Load a date even if it already exists in the ledger")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.export_file, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit("The export file looks empty.")

    header = rows[0]
    driver_col = 0  # column A is always "Driver"
    date_columns = []  # list of (col_index, resolved_date)
    for idx, col_header in enumerate(header):
        if idx == driver_col or col_header is None:
            continue
        resolved = parse_header_date(col_header, args.date)
        if resolved:
            date_columns.append((idx, resolved))

    if not date_columns:
        sys.exit(
            "Could not find a 'Total amount (date)' column, and no --date override "
            "was given. Pass --date YYYY-MM-DD to set it manually."
        )

    existing_dates = load_existing_dates(args.ledger)
    source_label = args.source_label or f"Business Office Export ({os.path.basename(args.export_file)})"

    new_rows = []
    for col_idx, as_of_date in date_columns:
        if as_of_date in existing_dates and not args.force:
            print(f"Skipping {as_of_date}: already present in {args.ledger} (use --force to reload it)")
            continue
        count = 0
        for row in rows[1:]:
            driver_text = row[driver_col]
            if driver_text is None:
                continue
            if str(driver_text).strip().lower() == "grand total":
                continue
            value = row[col_idx]
            if value is None or value == "":
                continue
            fund_code, fund_name, category = parse_driver(str(driver_text))
            new_rows.append([
                as_of_date, "", fund_code, fund_name, category,
                "Balance Reconciliation", "", value,
                "Weekly business office export", source_label,
            ])
            count += 1
        print(f"Prepared {count} fund rows for {as_of_date}")

    if not new_rows:
        print("Nothing new to add.")
        return

    file_exists = os.path.exists(args.ledger)
    os.makedirs(os.path.dirname(args.ledger) or ".", exist_ok=True)
    with open(args.ledger, "a", newline="") as f:
        w = csv.writer(f)
        if not file_exists:
            w.writerow(LEDGER_HEADER)
        w.writerows(new_rows)

    print(f"Appended {len(new_rows)} rows to {args.ledger}")


if __name__ == "__main__":
    main()
